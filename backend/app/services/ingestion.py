import os
from typing import List, Dict, Any, Optional
import pypdf
import docx
import openpyxl


class ParsedPage:
    def __init__(self, page_number: int, text: str):
        self.page_number = page_number
        self.text = text


class ChunkData:
    def __init__(self, content: str, page_number: Optional[int], token_count: int, chunk_index: int):
        self.content = content
        self.page_number = page_number
        self.token_count = token_count
        self.chunk_index = chunk_index


class DocumentIngestionService:
    @staticmethod
    def parse_file(file_path: str, file_type: str) -> List[ParsedPage]:
        """Parses PDF, DOCX, XLSX, or TXT files and returns a list of ParsedPage objects."""
        pages: List[ParsedPage] = []
        ext = file_type.lower().strip(".")

        if ext == "pdf":
            reader = pypdf.PdfReader(file_path)
            for idx, page in enumerate(reader.pages):
                text = page.extract_text() or ""
                if text.strip():
                    pages.append(ParsedPage(page_number=idx + 1, text=text))

        elif ext in ["docx", "doc"]:
            doc = docx.Document(file_path)
            full_text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
            # Word documents don't have explicit pagination, so treat as single page or split by ~1000 words
            pages.append(ParsedPage(page_number=1, text=full_text))

        elif ext in ["xlsx", "xls"]:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_texts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows_text = []
                for row in sheet.iter_rows(values_only=True):
                    row_str = " | ".join([str(val) for val in row if val is not None])
                    if row_str.strip():
                        rows_text.append(row_str)
                if rows_text:
                    sheet_texts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows_text))
            pages.append(ParsedPage(page_number=1, text="\n\n".join(sheet_texts)))

        elif ext in ["txt", "md", "csv"]:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
            pages.append(ParsedPage(page_number=1, text=text))

        else:
            raise ValueError(f"Unsupported file type: {file_type}")

        return pages

    @staticmethod
    def chunk_text(pages: List[ParsedPage], chunk_size_tokens: int = 400, overlap_tokens: int = 50) -> List[ChunkData]:
        """
        Splits text into chunks of roughly `chunk_size_tokens` (~4 chars per token)
        with `overlap_tokens` while preserving page_number metadata.
        """
        chunks: List[ChunkData] = []
        chunk_idx = 0

        # Rough conversion: 1 token ~= 4 characters
        chunk_char_size = chunk_size_tokens * 4
        overlap_char_size = overlap_tokens * 4

        for page in pages:
            text = page.text.strip()
            if not text:
                continue

            if len(text) <= chunk_char_size:
                token_count = max(1, len(text) // 4)
                chunks.append(ChunkData(
                    content=text,
                    page_number=page.page_number,
                    token_count=token_count,
                    chunk_index=chunk_idx
                ))
                chunk_idx += 1
            else:
                start = 0
                while start < len(text):
                    end = start + chunk_char_size
                    chunk_str = text[start:end]

                    # Try to break at a paragraph/newline boundary if available
                    if end < len(text):
                        last_newline = chunk_str.rfind("\n")
                        if last_newline > chunk_char_size // 2:
                            end = start + last_newline
                            chunk_str = text[start:end]

                    token_count = max(1, len(chunk_str) // 4)
                    chunks.append(ChunkData(
                        content=chunk_str.strip(),
                        page_number=page.page_number,
                        token_count=token_count,
                        chunk_index=chunk_idx
                    ))
                    chunk_idx += 1

                    if end >= len(text):
                        break

                    start = end - overlap_char_size

        return chunks
